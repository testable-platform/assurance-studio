#!/usr/bin/env python3
"""Build config/metrics_registry.yaml from Book.xlsx or White Box sheet (v0.2)."""

from __future__ import print_function

import os
import re
import sys

try:
    import yaml
except ImportError:
    yaml = None

import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
DEFAULT_BOOK = os.environ.get(
    "BOOK_XLSX",
    os.path.join(os.path.expanduser("~"), "Downloads", "Testable_Strategy_Metrics_Mapping_v0.2 (3).xlsx"),
)
OUT_YAML = os.path.join(ROOT, "config", "metrics_registry.yaml")
OUT_MD = os.path.join(ROOT, "docs", "METRICS_REGISTRY_SUMMARY.md")

TECHNIQUE_CODES = {
    ("Structural Analysis", "Cyclomatic Complexity"): "SA",
    ("Readability / Maintainability", "Cognitive Complexity"): "RM",
    ("Code Quality Auditing", "Code Duplication"): "CQ",
    ("Static Code Analysis", "Lint / Rule Violations"): "LR",
    ("Security White-box Testing", "Static Vulnerabilities (SAST)"): "SX",
    ("Security White-box Testing", "Dependency Risk (SCA)"): "DR",
    ("Control Flow Testing", "Statement Coverage"): "ST",
    ("Control Flow Testing", "Branch Coverage"): "BR",
    ("Control Flow Testing", "Path Coverage"): "PC",
    ("Mutation Testing", "Mutation Score"): "MU",
    ("Test Regression/Coverage Analysis", "Coverage Delta"): "CD",
    ("Data Flow Testing", "All Definition Coverage"): "DF",
    ("Data Flow Testing", "All Uses Coverage"): "DU",
    ("Development Process Analysis", "Code Churn"): "DP",
}

SA_METRIC_CODES = {
    "Execution Path Integrity": "EPI",
    "Decision Outcome Verification": "DOV",
    "Logical Sub-expression Validation": "LSV",
    "Total Logical Combinatorial Coverage": "TLCC",
    "Technical Debt Impact": "TDI",
    "QA Resource Allocation": "QRA",
}

LANG_COLS_BOOK = {
    "python": (7, 8),
    "c": (11, 12),
    "cpp": (13, 14),
    "java": (15, 16),
    "csharp": (19, 20),
    "javascript": (21, 22),
    "typescript": (23, 24),
}

STOP = {"and", "or", "the", "a", "an", "of", "per", "for", "to", "in", "on", "all", "uses", "use"}


def _slug_words(name):
    words = re.findall(r"[A-Za-z0-9]+", name)
    return [w for w in words if w.lower() not in STOP]


def _acronym(name, max_len=5):
    words = _slug_words(name)
    if not words:
        return "MET"
    if len(words) == 1:
        return words[0][:max_len].upper()
    code = "".join(w[0] for w in words).upper()
    if len(code) < 3:
        code = "".join(w[:2] for w in words).upper()
    return code[:max_len]


def _module_key(l5):
    return re.sub(r"[^a-z0-9]+", "_", l5.lower()).strip("_")[:64] or "metric"


def _branch_slug(l5):
    words = re.findall(r"[A-Za-z0-9]+", l5)
    return "-".join(w.capitalize() if w.islower() else w for w in words) or "Metric"


def _clean_l5(l4, l5):
    l5s = (l5 or "").strip()
    if not l5s or "Retrieving data" in l5s:
        return (l4 or "").strip()
    return l5s


def _norm_primary(val):
    if not val:
        return None
    return re.sub(r"\s+", " ", str(val).replace("\n", " ")).strip()


def _yes_no(val):
    if val is None:
        return False
    return str(val).strip().lower() in ("yes", "true", "1")


def parse_white_box(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["White Box"]
    techniques = {}
    metrics = []
    cur_l2 = cur_l3 = cur_l4 = ""

    for r in range(5, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 35)]
        if row[1]:
            cur_l2 = str(row[1]).strip()
        if row[2]:
            cur_l3 = str(row[2]).strip()
        if row[3]:
            cur_l4 = str(row[3]).strip()
        l5_raw = row[4]
        if not l5_raw:
            continue
        l5 = _clean_l5(cur_l4, l5_raw)
        if not l5 or not cur_l3:
            continue

        l2 = cur_l2 or "White Box"
        key = (l2, cur_l3)
        tech_code = TECHNIQUE_CODES.get(key)
        if not tech_code:
            raise ValueError("No technique_code for %s / %s (row %d)" % (l2, cur_l3, r))

        if tech_code not in techniques:
            techniques[tech_code] = {
                "technique_code": tech_code,
                "l2": l2,
                "l3": cur_l3,
                "report_group_label": l2,
                "metrics": [],
            }

        py_pri = _norm_primary(row[6])
        py_sec = _norm_primary(row[7])
        tools = {}
        if py_pri or py_sec:
            tools["python"] = {"primary": py_pri, "secondary": py_sec}

        metrics.append({
            "row": r,
            "technique_code": tech_code,
            "l4": cur_l4,
            "l5": l5,
            "tools": tools,
            "emitted_directly": _yes_no(row[8]),
            "derivation": str(row[9]).strip() if row[9] else "",
            "raw_formula": str(row[30]).strip() if row[30] else "",
            "expected_threshold": str(row[31]).strip() if row[31] else "",
            "normalisation": str(row[32]).strip() if row[32] else "",
            "execution_frequency": str(row[33]).strip() if row[33] else "",
        })

    return _finalize_techniques(techniques, metrics, os.path.basename(path))


def parse_book_sheet1(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    techniques = {}
    metrics = []
    for r in range(6, ws.max_row + 1):
        l2 = ws.cell(r, 2).value
        l3 = ws.cell(r, 3).value
        l4 = ws.cell(r, 4).value
        l5_raw = ws.cell(r, 5).value
        if not l2 or not l3:
            continue
        l2, l3 = str(l2).strip(), str(l3).strip()
        l4 = str(l4).strip() if l4 else ""
        l5 = _clean_l5(l4, l5_raw)
        if not l5:
            continue
        key = (l2, l3)
        tech_code = TECHNIQUE_CODES.get(key)
        if not tech_code:
            raise ValueError("No technique_code for %s / %s (row %d)" % (l2, l3, r))
        if tech_code not in techniques:
            techniques[tech_code] = {
                "technique_code": tech_code,
                "l2": l2,
                "l3": l3,
                "report_group_label": l2,
                "metrics": [],
            }
        tools = {}
        for lang, (pc, sc) in LANG_COLS_BOOK.items():
            pri, sec = ws.cell(r, pc).value, ws.cell(r, sc).value
            if pri or sec:
                tools[lang] = {
                    "primary": str(pri).strip() if pri else None,
                    "secondary": str(sec).strip() if sec else None,
                }
        metrics.append({
            "row": r,
            "technique_code": tech_code,
            "l4": l4,
            "l5": l5,
            "tools": tools,
            "emitted_directly": False,
            "derivation": "",
            "raw_formula": "",
            "expected_threshold": "",
            "normalisation": "",
            "execution_frequency": "",
        })
    return _finalize_techniques(techniques, metrics, os.path.basename(path))


def parse_workbook(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "White Box" in wb.sheetnames:
        wb.close()
        return parse_white_box(path)
    wb.close()
    return parse_book_sheet1(path)


def _finalize_techniques(techniques, metrics, source_name):
    for tech_code, tech in techniques.items():
        used = set()
        group_rows = [m for m in metrics if m["technique_code"] == tech_code]
        for m in group_rows:
            if tech_code == "SA" and m["l5"] in SA_METRIC_CODES:
                code = SA_METRIC_CODES[m["l5"]]
            else:
                code = _acronym(m["l5"])
            base = code
            n = 2
            while code in used:
                code = "%s%d" % (base[:4], n)
                n += 1
            used.add(code)
            mod_key = _module_key(m["l5"])
            entry = {
                "metric_code": code,
                "module_key": mod_key,
                "branch_slug": _branch_slug(m["l5"]),
                "l4_classification": m["l4"],
                "l5_metric": m["l5"],
                "taxonomy_classification": m["l4"],
                "tools": m["tools"],
                "emitted_directly": m.get("emitted_directly", False),
                "derivation": m.get("derivation", ""),
                "raw_formula": m.get("raw_formula", ""),
                "expected_threshold": m.get("expected_threshold", ""),
                "normalisation": m.get("normalisation", ""),
                "execution_frequency": m.get("execution_frequency", ""),
                "source_row": m["row"],
            }
            tech["metrics"].append(entry)

    return {
        "version": "3.1",
        "source": source_name,
        "branch_naming": "v2",
        "branch_types": ["Bug", "BugFX", "TCC", "CC"],
        "techniques": list(techniques.values()),
    }


def write_outputs(data):
    os.makedirs(os.path.dirname(OUT_YAML), exist_ok=True)
    os.makedirs(os.path.dirname(OUT_MD), exist_ok=True)
    if yaml is None:
        raise RuntimeError("PyYAML required: pip install pyyaml")
    with open(OUT_YAML, "w", encoding="utf-8") as fh:
        yaml.dump(data, fh, default_flow_style=False, sort_keys=False, allow_unicode=True)
    lines = [
        "# Metrics registry summary",
        "",
        "Parsed from **%s** — %d technique groups, %d L5 metrics." % (
            data["source"],
            len(data["techniques"]),
            sum(len(t["metrics"]) for t in data["techniques"]),
        ),
        "",
        "| technique | slug | module_key | L5 metric | Python primary | emitted |",
        "|-----------|------|------------|-----------|------------------|---------|",
    ]
    for tech in data["techniques"]:
        for m in tech["metrics"]:
            py = (m.get("tools") or {}).get("python", {})
            pri = py.get("primary") or "—"
            em = "Yes" if m.get("emitted_directly") else "No"
            lines.append("| %s | %s | %s | %s | %s | %s |" % (
                tech["technique_code"], m.get("branch_slug", ""), m["module_key"],
                m["l5_metric"], pri, em))
    with open(OUT_MD, "w", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_BOOK
    if not os.path.isfile(path):
        print("ERROR: workbook not found at %s" % path, file=sys.stderr)
        return 1
    data = parse_workbook(path)
    write_outputs(data)
    n_metrics = sum(len(t["metrics"]) for t in data["techniques"])
    print("Wrote %s (%d techniques, %d metrics)" % (OUT_YAML, len(data["techniques"]), n_metrics))
    print("Wrote %s" % OUT_MD)
    return 0


if __name__ == "__main__":
    sys.exit(main())
