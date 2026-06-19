"""Export comparison results to Excel."""

from __future__ import print_function

import io
from pathlib import Path

from openpyxl import Workbook


def _sheet_from_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def _flatten_diffs(branch, pair_name, pair):
    rows = []
    if not pair:
        return rows
    for diff in (pair.get("field_diffs") or []) + (pair.get("metric_diffs") or []):
        row = {
            "branch": branch,
            "pair": pair_name,
            "field": diff.get("field", ""),
            "match": diff.get("match"),
            "delta": diff.get("delta", ""),
        }
        for key, val in diff.items():
            if key not in ("field", "match", "delta"):
                row[key] = val
        rows.append(row)
    return rows


def build_comparison_workbook(results, whitebox_by_branch=None, path=None):
    """Build multi-sheet xlsx from comparison result dicts."""
    wb = Workbook()
    wb.remove(wb.active)
    whitebox_by_branch = whitebox_by_branch or {}

    summary_ws = wb.create_sheet("Summary")
    summary_headers = [
        "branch", "verdict", "taxonomy_status", "taxonomy_vs_s3",
        "s3_status", "local_status", "sonar_status",
        "s3_vs_local_verdict", "local_vs_sonar_verdict", "summary",
    ]
    summary_rows = []
    s3_local_rows = []
    local_sonar_rows = []
    status_rows = []
    failing_rows = []

    for row in results or []:
        branch = row.get("branch_name", "")
        wb_info = whitebox_by_branch.get(branch, {})
        summary_rows.append({
            "branch": branch,
            "verdict": row.get("verdict", ""),
            "taxonomy_status": row.get("taxonomy_status", ""),
            "taxonomy_vs_s3": row.get("taxonomy_vs_s3", ""),
            "s3_status": row.get("s3_status", ""),
            "local_status": row.get("local_status", ""),
            "sonar_status": row.get("sonar_status", ""),
            "s3_vs_local_verdict": (row.get("s3_vs_local") or {}).get("verdict", ""),
            "local_vs_sonar_verdict": (row.get("local_vs_sonar") or {}).get("verdict", ""),
            "summary": row.get("summary", ""),
        })
        status_rows.append({
            "branch": branch,
            "s3_status": row.get("s3_status", ""),
            "local_status": row.get("local_status", ""),
            "sonar_status": row.get("sonar_status", ""),
            "taxonomy_status_ref": row.get("taxonomy_status", ""),
            "taxonomy_vs_s3": row.get("taxonomy_vs_s3", ""),
            "run_status": wb_info.get("run_status", ""),
            "tasks": "%s/%s" % (
                wb_info.get("completed_tasks", 0), wb_info.get("total_tasks", 0),
            ) if wb_info.get("total_tasks") else "",
            "failed_tasks": wb_info.get("failed_tasks", ""),
            "run_health": wb_info.get("run_health", ""),
        })
        s3_local_rows.extend(_flatten_diffs(branch, "s3_vs_local", row.get("s3_vs_local")))
        local_sonar_rows.extend(_flatten_diffs(branch, "local_vs_sonar", row.get("local_vs_sonar")))
        for sec in wb_info.get("failing_sections") or []:
            failing_rows.append({
                "branch": branch,
                "testing_type": sec.get("testing_type", ""),
                "technique": sec.get("technique", ""),
                "classification": sec.get("classification", ""),
                "normalized_score": sec.get("normalized_score", ""),
            })

    _sheet_from_rows(summary_ws, summary_headers, summary_rows)

    status_ws = wb.create_sheet("Report status")
    _sheet_from_rows(
        status_ws,
        [
            "branch", "s3_status", "local_status", "sonar_status",
            "taxonomy_status_ref", "taxonomy_vs_s3",
            "run_status", "tasks", "failed_tasks", "run_health",
        ],
        status_rows,
    )

    s3_local_ws = wb.create_sheet("S3 vs Local")
    s3_headers = sorted({k for r in s3_local_rows for k in r}) if s3_local_rows else [
        "branch", "pair", "field", "match", "delta",
    ]
    _sheet_from_rows(s3_local_ws, s3_headers, s3_local_rows)

    ls_ws = wb.create_sheet("Local vs Sonar")
    ls_headers = sorted({k for r in local_sonar_rows for k in r}) if local_sonar_rows else [
        "branch", "pair", "field", "match", "delta",
    ]
    _sheet_from_rows(ls_ws, ls_headers, local_sonar_rows)

    fail_ws = wb.create_sheet("Failing sections")
    _sheet_from_rows(
        fail_ws,
        ["branch", "testing_type", "technique", "classification", "normalized_score"],
        failing_rows,
    )

    if path:
        wb.save(str(path))
        return str(path)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
