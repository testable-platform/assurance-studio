"""Tests for comparison semantics and export."""

from __future__ import print_function

import json
import os
import sys
import tempfile
import unittest

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, ROOT)

from io import BytesIO

from openpyxl import load_workbook

from lib.compare import compare_four_reports, compare_reports_pair  # noqa: E402
from lib.compare_export import build_comparison_workbook  # noqa: E402
from lib.proofs import collect_comparison_proof, whitebox_completion  # noqa: E402


def _report(branch, status, tool="Tool"):
    return {
        "branch_name": branch,
        "status": status,
        "tool_name": tool,
        "metric_values": {"score": 10.0},
    }


class RawVerdictTests(unittest.TestCase):
    def test_raw_verdict_matches_when_metrics_match_status_differs(self):
        s3 = _report("B_Bug_2.6", "PASS")
        local = _report("B_Bug_2.6", "FAIL")
        result = compare_reports_pair(s3, local, "s3", "local", raw_verdict=True, shared_metrics_only=True)
        self.assertEqual(result["verdict"], "MATCH")
        self.assertFalse(result["status_match"])

    def test_raw_verdict_mismatch_when_metrics_differ(self):
        s3 = _report("B_Bug_2.6", "PASS")
        local = dict(_report("B_Bug_2.6", "PASS"))
        local["metric_values"] = {"score": 99.0}
        result = compare_reports_pair(s3, local, "s3", "local", raw_verdict=True, shared_metrics_only=True)
        self.assertEqual(result["verdict"], "MISMATCH")

    def test_raw_verdict_incomplete_without_shared_metrics(self):
        s3 = _report("B_Bug_2.6", "PASS")
        local = dict(_report("B_Bug_2.6", "PASS"))
        local["metric_values"] = {"other": 1}
        result = compare_reports_pair(s3, local, "s3", "local", raw_verdict=True, shared_metrics_only=True)
        self.assertEqual(result["verdict"], "INCOMPLETE")


class CompareSemanticsTests(unittest.TestCase):
    def test_verdict_ignores_taxonomy_status(self):
        tax = _report("B_Bug_2.6", "FAIL")
        s3 = _report("B_Bug_2.6", "PASS")
        local = _report("B_Bug_2.6", "PASS")
        sonar = {"branch_name": "B_Bug_2.6", "status": "SKIPPED"}
        result = compare_four_reports(tax, s3, local, sonar)
        self.assertIn(result["verdict"], ("MATCH", "PARTIAL"))
        self.assertEqual(result["taxonomy_vs_s3"], "DIFFERS")

    def test_s3_na_and_local_pass_is_partial(self):
        branch = "SA_Decision-Outcome-Verification_Bug_2.6"
        tax = _report(branch, "PASS")
        s3 = _report(branch, "N/A")
        local = _report(branch, "PASS", "Coverage.py")
        sonar = {"branch_name": branch, "status": "SKIPPED"}
        result = compare_four_reports(tax, s3, local, sonar)
        self.assertEqual(result["verdict"], "PARTIAL")
        self.assertEqual(result["s3_vs_local"]["verdict"], "N/A")
        self.assertEqual(result["taxonomy_vs_local"], "AGREE")
        self.assertEqual(result["taxonomy_vs_s3"], "N/A")

    def test_bugfx_taxonomy_local_agree(self):
        branch = "SA_Decision-Outcome-Verification_BugFX_2.6"
        tax = _report(branch, "PASS")
        s3 = _report(branch, "N/A")
        local = _report(branch, "PASS", "Coverage.py")
        sonar = {"branch_name": branch, "status": "SKIPPED"}
        result = compare_four_reports(tax, s3, local, sonar)
        self.assertEqual(result["taxonomy_vs_local"], "AGREE")
        self.assertEqual(result["local_status"], "PASS")

    def test_local_only_comparison(self):
        tmp = tempfile.mkdtemp()
        proof = os.path.join(tmp, "proofs", "DF", "DF_DU-Path-Validation_Bug_2.6")
        os.makedirs(proof, exist_ok=True)
        local = _report("DF_DU-Path-Validation_Bug_2.6", "PASS", "Beniget")
        with open(os.path.join(proof, "local_report.json"), "w", encoding="utf-8") as fh:
            json.dump(local, fh)
        result = collect_comparison_proof("DF_DU-Path-Validation_Bug_2.6", root=tmp)
        self.assertEqual(result["verdict"], "PARTIAL")
        self.assertIn("s3_report.json", result.get("missing_reports", []))

    def test_excel_workbook_builds(self):
        results = [{
            "branch_name": "DF_DU-Path-Validation_Bug_2.6",
            "verdict": "PARTIAL",
            "taxonomy_status": "PASS",
            "taxonomy_vs_s3": "N/A",
            "taxonomy_vs_local": "AGREE",
            "s3_status": "SKIPPED",
            "local_status": "PASS",
            "sonar_status": "SKIPPED",
            "local_tool_name": "Beniget",
            "local_executed_tool": "beniget",
            "local_real_tool": True,
            "local_metric_values": {"score": 10.0},
            "s3_vs_local": {"verdict": "N/A", "field_diffs": [], "metric_diffs": []},
            "local_vs_sonar": {"verdict": "N/A", "field_diffs": [], "metric_diffs": []},
            "summary": "test",
        }]
        data = build_comparison_workbook(results)
        self.assertTrue(len(data) > 1000)

    def test_excel_workbook_has_matches_sheet(self):
        results = [{
            "branch_name": "DF_DU-Path-Validation_Bug_2.6",
            "verdict": "MATCH",
            "taxonomy_status": "PASS",
            "taxonomy_vs_s3": "AGREE",
            "taxonomy_vs_local": "AGREE",
            "s3_status": "PASS",
            "local_status": "PASS",
            "sonar_status": "SKIPPED",
            "local_tool_name": "Beniget",
            "local_metric_values": {"score": 10.0},
            "s3_vs_local": {
                "verdict": "MATCH",
                "field_diffs": [
                    {"field": "status", "s3": "PASS", "local": "PASS", "match": True},
                ],
                "metric_diffs": [
                    {"field": "score", "s3": 10.0, "local": 10.0, "match": True, "delta": 0.0},
                ],
            },
            "local_vs_sonar": {"verdict": "N/A", "field_diffs": [], "metric_diffs": []},
            "summary": "aligned",
        }]
        data = build_comparison_workbook(results)
        wb = load_workbook(BytesIO(data))
        self.assertIn("Matches", wb.sheetnames)
        self.assertIn("Mismatches", wb.sheetnames)
        self.assertIn("Local results", wb.sheetnames)
        matches = list(wb["Matches"].iter_rows(min_row=2, values_only=True))
        self.assertGreaterEqual(len(matches), 1)
        summary_row = list(wb["Summary"].iter_rows(min_row=2, values_only=True))[0]
        headers = [cell.value for cell in wb["Summary"][1]]
        self.assertIn("taxonomy_vs_local", headers)
        self.assertIn("matched_fields", headers)
        matched_idx = headers.index("matched_fields")
        self.assertGreater(summary_row[matched_idx], 0)

    def test_excel_mismatch_columns_and_detail(self):
        results = [{
            "branch_name": "SA_Decision-Outcome-Verification_BugFX_2.6",
            "verdict": "MISMATCH",
            "taxonomy_status": "PASS",
            "taxonomy_vs_s3": "DIFFERS",
            "taxonomy_vs_local": "AGREE",
            "s3_status": "PASS",
            "local_status": "FAIL",
            "sonar_status": "SKIPPED",
            "s3_tool_name": "Coverage.py",
            "s3_executed": True,
            "local_tool_name": "Coverage.py",
            "local_real_tool": True,
            "local_executed_tool": "coverage.py",
            "s3_vs_local": {
                "verdict": "MISMATCH",
                "field_diffs": [
                    {"field": "status", "s3": "PASS", "local": "FAIL", "match": False},
                ],
                "metric_diffs": [
                    {"field": "score", "s3": 10.0, "local": 99.0, "match": False, "delta": 89.0},
                ],
            },
            "local_vs_sonar": {"verdict": "N/A", "field_diffs": [], "metric_diffs": []},
            "summary": "score differs",
        }]
        data = build_comparison_workbook(results)
        wb = load_workbook(BytesIO(data))
        mismatch_headers = [cell.value for cell in wb["Mismatches"][1]]
        for col in (
            "mismatched_field", "S3 Data", "Local Data", "S3 vs Local",
            "S3 tool executed", "Local tool executed",
        ):
            self.assertIn(col, mismatch_headers)
        mismatch_rows = list(wb["Mismatches"].iter_rows(min_row=2, values_only=True))
        self.assertGreaterEqual(len(mismatch_rows), 1)
        status_idx = mismatch_headers.index("S3 vs Local")
        self.assertEqual(mismatch_rows[0][status_idx], "MISMATCH")
        summary_headers = [cell.value for cell in wb["Summary"][1]]
        self.assertIn("mismatched_fields_detail", summary_headers)
        self.assertIn("S3 Data", summary_headers)
        self.assertIn("Local Data", summary_headers)
        detail_idx = summary_headers.index("mismatched_fields_detail")
        summary_row = list(wb["Summary"].iter_rows(min_row=2, values_only=True))[0]
        self.assertIn("score", summary_row[detail_idx])


class LocalProofSmokeTests(unittest.TestCase):
    _branch_dir = os.path.join(
        ROOT, ".pipeline_work", "08b2473015be749b", "SA_Decision-Outcome-Verification_Bug_2.6",
    )

    @unittest.skipUnless(os.path.isdir(_branch_dir), "pipeline branch copy not present")
    def test_collect_local_proof_smoke(self):
        from lib.proofs import collect_local_proof

        tmp = tempfile.mkdtemp()
        work = os.path.join(tmp, "work")
        os.makedirs(work, exist_ok=True)
        import shutil
        shutil.copytree(self._branch_dir, os.path.join(work, "SA_Decision-Outcome-Verification_Bug_2.6"))
        try:
            report = collect_local_proof(
                "SA_Decision-Outcome-Verification_Bug_2.6",
                install=False,
                root=tmp,
                local_root=work,
            )
        except Exception as exc:
            self.skipTest("local tool env unavailable: %s" % exc)
        self.assertIn(report.get("status"), ("PASS", "FAIL", "WARN", "ERROR", "UNAVAILABLE"))
        extra = report.get("extra") or {}
        if report.get("status") in ("PASS", "FAIL", "WARN"):
            self.assertTrue(extra.get("real_tool", True) is not False)


class WhiteboxRunHealthTests(unittest.TestCase):
    def test_degraded_run_health_from_summary(self):
        tmp = tempfile.mkdtemp()
        tax_root = os.path.join(tmp, "taxonomy_reports", "Data Flow Testing")
        branch = "DF_DU-Path-Validation_Bug_2.6"
        folder = os.path.join(tax_root, "%s_20260618T120000Z" % branch)
        os.makedirs(folder, exist_ok=True)
        html = os.path.join(folder, "taxonomy-gate-run-001.html")
        with open(html, "w", encoding="utf-8") as fh:
            fh.write(
                "Branch name</th><td>%s</td>Commit ID</th><td>abc</td>Run ID</th><td>run-001</td>"
                % branch
            )
        with open(os.path.join(folder, "run_summary.json"), "w", encoding="utf-8") as fh:
            json.dump({
                "status": "failed",
                "total_tasks": 28,
                "completed_tasks": 27,
                "failed_tasks": 1,
            }, fh)
        result = whitebox_completion([branch], root=tmp)
        info = result[branch]
        self.assertEqual(info["status"], "COMPLETED")
        self.assertEqual(info["run_health"], "OK")
        self.assertEqual(info["failed_tasks"], 1)


if __name__ == "__main__":
    unittest.main()
